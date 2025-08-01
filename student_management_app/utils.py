import math
from geopy.distance import geodesic
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
import os
from datetime import datetime
from django.conf import settings
import tempfile

def calculate_distance(lat1, lon1, lat2, lon2, accuracy1=None, accuracy2=None):
    """
    Calculate the distance between two geographic coordinates using geodesic distance.
    Returns distance in meters.

    Parameters:
    - lat1, lon1: First point coordinates
    - lat2, lon2: Second point coordinates
    - accuracy1: Optional accuracy in meters for the first point
    - accuracy2: Optional accuracy in meters for the second point
    """
    if lat1 is None or lon1 is None or lat2 is None or lon2 is None:
        return {
            'distance': float('inf'),  # Return infinity if any coordinate is None
            'error_margin': float('inf'),
            'is_reliable': False
        }

    try:
        # Use geopy's geodesic distance calculation (more accurate than manual calculation)
        point1 = (float(lat1), float(lon1))
        point2 = (float(lat2), float(lon2))

        # Calculate distance in meters
        distance = geodesic(point1, point2).meters

        # Calculate error margin based on provided accuracy values
        error_margin = 0
        if accuracy1 is not None:
            error_margin += float(accuracy1)
        if accuracy2 is not None:
            error_margin += float(accuracy2)

        # If no accuracy provided, use a default error margin of 10 meters
        if error_margin == 0:
            error_margin = 10

        # Determine if the measurement is reliable
        # A measurement is considered reliable if the error margin is less than 50% of the distance
        is_reliable = error_margin < (distance * 0.5)

        return {
            'distance': distance,
            'error_margin': error_margin,
            'is_reliable': is_reliable
        }
    except Exception as e:
        # Return error information if calculation fails
        return {
            'distance': float('inf'),
            'error_margin': float('inf'),
            'is_reliable': False,
            'error': str(e)
        }

def is_within_radius(student_lat, student_lon, teacher_lat, teacher_lon, radius, student_accuracy=None, teacher_accuracy=None):
    """
    Check if the student is within the allowed radius of the teacher's location.
    Returns a dictionary with verification details.

    Parameters:
    - student_lat, student_lon: Student's coordinates
    - teacher_lat, teacher_lon: Teacher's coordinates
    - radius: Allowed radius in meters
    - student_accuracy: Optional accuracy of student's location in meters
    - teacher_accuracy: Optional accuracy of teacher's location in meters
    """
    # Calculate distance with accuracy information
    result = calculate_distance(
        student_lat, student_lon,
        teacher_lat, teacher_lon,
        student_accuracy, teacher_accuracy
    )

    # Get the distance from the result
    distance = result['distance']
    error_margin = result['error_margin']
    is_reliable = result['is_reliable']

    # Adjust the effective radius based on the error margin
    effective_radius = float(radius)

    # If the measurement is not reliable but we still want to verify,
    # we can reduce the effective radius to account for the uncertainty
    if not is_reliable and error_margin < float(radius):
        effective_radius = max(float(radius) - error_margin, 0)

    # Determine if the student is within the radius
    is_within = distance <= effective_radius

    # Return comprehensive verification result with proper boolean values
    return {
        'is_within': bool(is_within),  # Ensure it's a proper boolean
        'distance': float(distance),   # Ensure it's a proper float
        'error_margin': float(error_margin),
        'is_reliable': bool(is_reliable),  # Ensure it's a proper boolean
        'effective_radius': float(effective_radius),
        'original_radius': float(radius)
    }


def get_client_ip(request):
    """
    Get the client's IP address from the request.
    Handles cases where the request comes through proxies.
    """
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0].strip()
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def is_same_network(ip1, ip2, subnet_mask='255.255.255.0'):
    """
    Check if two IP addresses are on the same network.
    Default subnet mask assumes a /24 network (255.255.255.0).

    Parameters:
    - ip1, ip2: IP addresses to compare
    - subnet_mask: Network subnet mask (default: 255.255.255.0 for /24)

    Returns:
    - bool: True if IPs are on the same network, False otherwise
    """
    import ipaddress

    try:
        # Convert IPs to IPv4Address objects
        addr1 = ipaddress.IPv4Address(ip1)
        addr2 = ipaddress.IPv4Address(ip2)

        # Create network from first IP with /24 subnet (most common for local networks)
        network1 = ipaddress.IPv4Network(f"{ip1}/24", strict=False)

        # Check if second IP is in the same network
        return addr2 in network1

    except (ipaddress.AddressValueError, ValueError) as e:
        print(f"Error comparing networks: {e}")
        return False


def verify_network_connectivity(student_ip, teacher_ip, student_ssid=None, teacher_ssid=None):
    """
    Verify if student and teacher are on the same network.

    Parameters:
    - student_ip: Student's IP address
    - teacher_ip: Teacher's IP address
    - student_ssid: Student's WiFi network name (optional)
    - teacher_ssid: Teacher's WiFi network name (optional)

    Returns:
    - dict: Network verification results
    """
    verification_result = {
        'is_same_network': False,
        'ip_match': False,
        'ssid_match': False,
        'verification_method': None,
        'details': {}
    }

    # Check IP network match
    if student_ip and teacher_ip:
        ip_same_network = is_same_network(student_ip, teacher_ip)
        verification_result['ip_match'] = ip_same_network
        verification_result['details']['student_ip'] = student_ip
        verification_result['details']['teacher_ip'] = teacher_ip

        if ip_same_network:
            verification_result['is_same_network'] = True
            verification_result['verification_method'] = 'ip_network'

    # Check SSID match (if available)
    if student_ssid and teacher_ssid:
        ssid_match = student_ssid.lower() == teacher_ssid.lower()
        verification_result['ssid_match'] = ssid_match
        verification_result['details']['student_ssid'] = student_ssid
        verification_result['details']['teacher_ssid'] = teacher_ssid

        if ssid_match:
            verification_result['is_same_network'] = True
            verification_result['verification_method'] = 'wifi_ssid'

    # If both methods are available, use OR logic (either one can pass) for better usability
    if student_ssid and teacher_ssid and student_ip and teacher_ip:
        # Use OR logic - if either IP or SSID matches, consider it same network
        verification_result['is_same_network'] = verification_result['ip_match'] or verification_result['ssid_match']
        if verification_result['ip_match'] and verification_result['ssid_match']:
            verification_result['verification_method'] = 'both_ip_and_ssid'
        elif verification_result['ip_match']:
            verification_result['verification_method'] = 'ip_network'
        elif verification_result['ssid_match']:
            verification_result['verification_method'] = 'wifi_ssid'

    # Add debugging information
    print(f"Network verification debug:")
    print(f"- Student IP: {student_ip}")
    print(f"- Teacher IP: {teacher_ip}")
    print(f"- Student SSID: {student_ssid}")
    print(f"- Teacher SSID: {teacher_ssid}")
    print(f"- IP Match: {verification_result['ip_match']}")
    print(f"- SSID Match: {verification_result['ssid_match']}")
    print(f"- Final Result: {verification_result['is_same_network']}")
    print(f"- Method: {verification_result['verification_method']}")

    return verification_result

def export_attendance_to_excel(attendance_data, subject_name=None, date_range=None, for_student=False):
    """
    Generate Excel file from attendance data

    Parameters:
    - attendance_data: QuerySet of AttendanceReport objects
    - subject_name: Optional subject name for the report title
    - date_range: Optional date range for the report title
    - for_student: Boolean indicating if this is a student's personal report

    Returns:
    - Path to the temporary Excel file
    """
    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Set the title of the worksheet
    title = "Attendance Report"
    if subject_name:
        title += f" - {subject_name}"
    if date_range:
        title += f" ({date_range})"

    worksheet.title = "Attendance Data"

    # Define styles
    header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Only add title and info for student reports, not for teacher reports (import compatibility)
    if for_student:
        # Add title
        worksheet.merge_cells('A1:F1')
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = Font(name='Arial', bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Add report generation info
        worksheet.merge_cells('A2:F2')
        gen_info_cell = worksheet.cell(row=2, column=1)
        gen_info_cell.value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        gen_info_cell.font = Font(italic=True)
        gen_info_cell.alignment = Alignment(horizontal='center')

    # Define headers based on report type
    if for_student:
        headers = ['Date', 'Subject', 'Status', 'Location Verified', 'Teacher']
        # Add headers at row 4 for student reports (keep title and info)
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        row_num = 5
    else:
        # Use exact column names expected by the import function - keep it simple for import compatibility
        headers = ['Student ID', 'Student Name', 'Date', 'Status']

        # For import compatibility, start headers at row 1 (no title or merged cells)
        # Add headers directly at row 1
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add data rows starting from row 2
        row_num = 2
    present_count = 0
    absent_count = 0
    location_verified_count = 0

    for record in attendance_data:
        if for_student:
            # Student's personal report
            worksheet.cell(row=row_num, column=1).value = record.attendance_id.attendance_date.strftime('%Y-%m-%d')
            worksheet.cell(row=row_num, column=2).value = record.attendance_id.subject_id.subject_name
            status = "Present" if record.status else "Absent"
            worksheet.cell(row=row_num, column=3).value = status
            location_verified = "Yes" if hasattr(record, 'location_verified') and record.location_verified else "No"
            worksheet.cell(row=row_num, column=4).value = location_verified
            worksheet.cell(row=row_num, column=5).value = f"{record.attendance_id.subject_id.staff_id.admin.first_name} {record.attendance_id.subject_id.staff_id.admin.last_name}"
        else:
            # Teacher's report for all students - simplified for easy import
            # Column 1: Student ID (username for import compatibility)
            worksheet.cell(row=row_num, column=1).value = record.student_id.admin.username
            # Column 2: Student Name (for reference only)
            worksheet.cell(row=row_num, column=2).value = f"{record.student_id.admin.first_name} {record.student_id.admin.last_name}"
            # Column 3: Date
            worksheet.cell(row=row_num, column=3).value = record.attendance_id.attendance_date.strftime('%Y-%m-%d')
            # Column 4: Status (Present/Absent)
            status = "Present" if record.status else "Absent"
            worksheet.cell(row=row_num, column=4).value = status

        # Update statistics
        if record.status:
            present_count += 1
        else:
            absent_count += 1

        if hasattr(record, 'location_verified') and record.location_verified:
            location_verified_count += 1

        row_num += 1

    # Only add statistics for student reports, not for teacher reports (import compatibility)
    if for_student:
        # Add statistics section
        row_num += 2  # Add some space

        worksheet.cell(row=row_num, column=1).value = "Statistics"
        worksheet.cell(row=row_num, column=1).font = Font(bold=True)

        worksheet.cell(row=row_num+1, column=1).value = "Total Records:"
        worksheet.cell(row=row_num+1, column=2).value = len(attendance_data)

        worksheet.cell(row=row_num+2, column=1).value = "Present Count:"
        worksheet.cell(row=row_num+2, column=2).value = present_count

        worksheet.cell(row=row_num+3, column=1).value = "Absent Count:"
        worksheet.cell(row=row_num+3, column=2).value = absent_count

        worksheet.cell(row=row_num+4, column=1).value = "Location Verified Count:"
        worksheet.cell(row=row_num+4, column=2).value = location_verified_count

        if present_count > 0:
            location_verification_rate = (location_verified_count / present_count) * 100
            worksheet.cell(row=row_num+5, column=1).value = "Location Verification Rate:"
            worksheet.cell(row=row_num+5, column=2).value = f"{location_verification_rate:.2f}%"

    # Auto-adjust column widths - Fixed to handle merged cells
    for i, column in enumerate(worksheet.columns, 1):
        length = 0
        for cell in column:
            try:
                if cell.value:
                    length = max(length, len(str(cell.value)))
            except:
                pass  # Skip cells that don't have a value attribute (like merged cells)

        if length > 0:
            # Get column letter safely using column index
            column_letter = openpyxl.utils.get_column_letter(i)
            # Setting width with a little buffer
            worksheet.column_dimensions[column_letter].width = length + 2

    # Create a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', prefix='attendance_export_') as tmp:
        temp_filename = tmp.name
        # Close the file handle before saving to avoid file access issues on Windows

    # Save workbook to the closed temp file
    try:
        workbook.save(temp_filename)
    except Exception as e:
        # If saving fails, make sure we clean up
        if os.path.exists(temp_filename):
            try:
                os.unlink(temp_filename)
            except:
                pass
        raise e

    return temp_filename

def export_attendance_to_excel_new(attendance_reports, subject_name="", date_range="", for_student=False):
    """
    Export attendance data to Excel

    Parameters:
    - attendance_reports: QuerySet of AttendanceReport objects
    - subject_name: Name of the subject (optional)
    - date_range: String representing the date range (optional)
    - for_student: Boolean indicating if the export is for a student (vs. teacher)

    Returns:
    - Path to the generated Excel file
    """
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Define styles
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Define borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Add title
    title = f"Attendance Report - {subject_name}"
    if date_range:
        title += f" ({date_range})"

    ws.merge_cells('A1:G1')
    ws['A1'] = title
    ws['A1'].font = Font(name='Arial', size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Add header row
    headers = ['Date', 'Subject', 'Student Name']

    if not for_student:
        headers.append('Student ID')

    headers.extend(['Status', 'Location Verified'])

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data rows
    row_num = 4

    # Group reports by date and subject for better organization
    if for_student:
        # For student report, group by date and subject
        reports_by_date = {}
        for report in attendance_reports:
            date_key = report.attendance_id.attendance_date
            subject_key = report.attendance_id.subject_id.subject_name

            if date_key not in reports_by_date:
                reports_by_date[date_key] = {}

            if subject_key not in reports_by_date[date_key]:
                reports_by_date[date_key][subject_key] = []

            reports_by_date[date_key][subject_key].append(report)

        # Sort by date
        for date in sorted(reports_by_date.keys()):
            for subject in sorted(reports_by_date[date].keys()):
                for report in reports_by_date[date][subject]:
                    # Date
                    cell = ws.cell(row=row_num, column=1)
                    cell.value = report.attendance_id.attendance_date.strftime('%Y-%m-%d')
                    cell.border = thin_border

                    # Subject
                    cell = ws.cell(row=row_num, column=2)
                    cell.value = report.attendance_id.subject_id.subject_name
                    cell.border = thin_border

                    # Student Name
                    cell = ws.cell(row=row_num, column=3)
                    cell.value = f"{report.student_id.admin.first_name} {report.student_id.admin.last_name}"
                    cell.border = thin_border

                    # Status
                    cell = ws.cell(row=row_num, column=4)
                    cell.value = "Present" if report.status else "Absent"
                    cell.border = thin_border

                    # Location Verified
                    cell = ws.cell(row=row_num, column=5)
                    cell.value = "Yes" if report.location_verified else "No"
                    cell.border = thin_border

                    row_num += 1
    else:
        # For teacher report, group by date, then by student
        reports_by_date = {}
        for report in attendance_reports:
            date_key = report.attendance_id.attendance_date

            if date_key not in reports_by_date:
                reports_by_date[date_key] = []

            reports_by_date[date_key].append(report)

        # Sort by date
        for date in sorted(reports_by_date.keys()):
            # Sort reports by student name
            sorted_reports = sorted(
                reports_by_date[date],
                key=lambda r: f"{r.student_id.admin.first_name} {r.student_id.admin.last_name}"
            )

            for report in sorted_reports:
                # Date
                cell = ws.cell(row=row_num, column=1)
                cell.value = report.attendance_id.attendance_date.strftime('%Y-%m-%d')
                cell.border = thin_border

                # Subject
                cell = ws.cell(row=row_num, column=2)
                cell.value = report.attendance_id.subject_id.subject_name
                cell.border = thin_border

                # Student Name
                cell = ws.cell(row=row_num, column=3)
                cell.value = f"{report.student_id.admin.first_name} {report.student_id.admin.last_name}"
                cell.border = thin_border

                # Student ID
                cell = ws.cell(row=row_num, column=4)
                cell.value = report.student_id.admin.username
                cell.border = thin_border

                # Status
                cell = ws.cell(row=row_num, column=5)
                cell.value = "Present" if report.status else "Absent"
                cell.border = thin_border

                # Location Verified
                cell = ws.cell(row=row_num, column=6)
                cell.value = "Yes" if report.location_verified else "No"
                cell.border = thin_border

                row_num += 1

    # Add summary at the bottom
    row_num += 2
    ws.cell(row=row_num, column=1).value = "Summary"
    ws.cell(row=row_num, column=1).font = Font(bold=True)

    row_num += 1
    total_present = sum(1 for report in attendance_reports if report.status)
    total_absent = sum(1 for report in attendance_reports if not report.status)
    total_records = len(attendance_reports)

    ws.cell(row=row_num, column=1).value = "Total Present:"
    ws.cell(row=row_num, column=2).value = total_present

    row_num += 1
    ws.cell(row=row_num, column=1).value = "Total Absent:"
    ws.cell(row=row_num, column=2).value = total_absent

    row_num += 1
    ws.cell(row=row_num, column=1).value = "Total Records:"
    ws.cell(row=row_num, column=2).value = total_records

    if total_records > 0:
        row_num += 1
        attendance_percentage = (total_present / total_records) * 100
        ws.cell(row=row_num, column=1).value = "Attendance Percentage:"
        ws.cell(row=row_num, column=2).value = f"{attendance_percentage:.2f}%"

    # Add timestamp
    row_num += 2
    ws.cell(row=row_num, column=1).value = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    # Auto-adjust column widths
    for i, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(i)

        for cell in column:
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create a temporary file
    fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
    os.close(fd)

    # Save the workbook to the temporary file
    wb.save(temp_path)

    return temp_path
