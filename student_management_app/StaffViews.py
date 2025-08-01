from django.shortcuts import render, redirect
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse, FileResponse
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.core import serializers
from django.utils.timezone import now
import json
import random
import string
import datetime
import uuid
import qrcode
import os
import tempfile
from io import BytesIO
from django.core.files.base import ContentFile
from django.core.cache import cache
import openpyxl
import openpyxl.styles

# Try to import pandas, but handle the case where it's not installed
try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except (ImportError, AttributeError):
    PANDAS_AVAILABLE = False

from student_management_app.models import (
    CustomUser, Staffs, Courses, Subjects, Students,
    SessionYearModel, Attendance, AttendanceReport, StudentResult, AttendanceQRCode
)
from .utils import get_client_ip, verify_network_connectivity
from .utils import export_attendance_to_excel

def staff_generate_qr(request):
    print(f"QR Generation request: {request.method}")  # Debug
    if request.method == "POST":
        print(f"POST data: {request.POST}")  # Debug
        subject_id = request.POST.get('subject')
        session_year_id = request.POST.get('session_year')
        expiry_minutes = request.POST.get('expiry_time', 30)
        teacher_latitude = request.POST.get('latitude')
        teacher_longitude = request.POST.get('longitude')
        allowed_radius = request.POST.get('radius', 100)

        # Validate and sanitize radius value
        try:
            allowed_radius = float(allowed_radius)
            # For testing purposes, allow larger radius (up to 50km)
            # In production, you might want to limit this to 1000m
            if allowed_radius < 10:
                allowed_radius = 10
            elif allowed_radius > 50000:  # 50km for testing
                allowed_radius = 50000
        except (ValueError, TypeError):
            allowed_radius = 100  # Default fallback

        # Get teacher's network information
        teacher_ip = get_client_ip(request)
        enable_network_verification = request.POST.get('enableNetwork') == 'on'

        print(f"Network verification settings:")
        print(f"- Teacher IP: {teacher_ip}")
        print(f"- Enable network verification: {enable_network_verification}")

        print(f"Parsed data - Subject: {subject_id}, Session: {session_year_id}, Expiry: {expiry_minutes}")  # Debug

        # Ensure required fields are present
        if not subject_id or not session_year_id:
            print("Missing required fields!")  # Debug
            return JsonResponse({"status": "error", "message": "Subject and session year are required."}, status=400)

        try:
            print(f"Looking for subject with ID: {subject_id}")  # Debug
            subject = Subjects.objects.get(id=subject_id)
            print(f"Found subject: {subject.subject_name}")  # Debug

            print(f"Looking for session year with ID: {session_year_id}")  # Debug
            session_year = SessionYearModel.objects.get(id=session_year_id)
            print(f"Found session year: {session_year}")  # Debug

            unique_token = str(uuid.uuid4())
            print(f"Generated token: {unique_token}")  # Debug

            # Create a URL that includes the token for direct scanning
            # This URL will redirect to the login page if user is not logged in
            qr_data = f"{request.build_absolute_uri('/scan-attendance/')}?token={unique_token}"
            print(f"QR data URL: {qr_data}")  # Debug

            print("Creating QR code...")  # Debug
            qr = qrcode.make(qr_data)
            qr_io = BytesIO()
            qr.save(qr_io, format="PNG")
            print("QR code created successfully")  # Debug

            # Create QR code instance with basic fields first
            print("Creating QR code instance...")  # Debug
            qr_code_instance = AttendanceQRCode(
                subject=subject,
                session_year=session_year,
                expiry_time=now() + datetime.timedelta(minutes=int(expiry_minutes)),
                is_active=True,
                token=unique_token,
                teacher_latitude=float(teacher_latitude) if teacher_latitude else None,
                teacher_longitude=float(teacher_longitude) if teacher_longitude else None,
                allowed_radius=float(allowed_radius)
            )

            # Store network information in cache using the token as key
            # This will expire when the QR code expires
            if enable_network_verification:
                network_info = {
                    'teacher_ip': teacher_ip,
                    'teacher_ssid': None,  # IP-based verification only
                    'require_network_verification': True
                }
                cache_key = f"qr_network_{unique_token}"
                cache.set(cache_key, network_info, timeout=int(expiry_minutes) * 60)  # Cache for same duration as QR code
                print(f"Stored network info in cache: {network_info}")  # Debug
            print("QR code instance created")  # Debug

            # Network verification removed
            print("Saving QR code image...")  # Debug
            qr_code_instance.qr_code_image.save(f"qr_{subject.id}_{session_year.id}.png", ContentFile(qr_io.getvalue()), save=True)
            print("QR code image saved successfully")  # Debug

            # Get the full URL to the QR code image
            qr_code_url = qr_code_instance.qr_code_image.url
            print(f"QR code URL: {qr_code_url}")  # Debug

            # Also include the base64 encoded image data for direct sharing
            import base64
            qr_io.seek(0)  # Reset the pointer to the beginning of the BytesIO object
            qr_base64 = base64.b64encode(qr_io.getvalue()).decode('utf-8')
            qr_data_url = f"data:image/png;base64,{qr_base64}"
            print("Base64 encoding completed")  # Debug

            response_data = {
                "status": "success",
                "qr_code_url": qr_code_instance.qr_code_image.url,
                "qr_data_url": qr_data_url,  # Include the base64 data URL
                "expiry_time": qr_code_instance.expiry_time.strftime("%Y-%m-%d %H:%M:%S"),
            }
            print(f"Returning success response: {response_data}")  # Debug
            return JsonResponse(response_data)

        except Subjects.DoesNotExist:
            print(f"Subject not found with ID: {subject_id}")  # Debug
            return JsonResponse({"status": "error", "message": "Invalid subject ID."}, status=400)
        except SessionYearModel.DoesNotExist:
            print(f"Session year not found with ID: {session_year_id}")  # Debug
            return JsonResponse({"status": "error", "message": "Invalid session year ID."}, status=400)
        except Exception as e:
            print(f"Unexpected error in QR generation: {str(e)}")  # Debug
            import traceback
            traceback.print_exc()  # Print full traceback for debugging
            return JsonResponse({"status": "error", "message": f"Error generating QR code: {str(e)}"}, status=400)

    return JsonResponse({"status": "error", "message": "Invalid request method."}, status=400)



def staff_home(request):
    # Get the Staff instance linked to the logged-in user
    staff_instance = Staffs.objects.get(admin=request.user)

    # Fetch only subjects assigned to this staff
    subjects = Subjects.objects.filter(staff_id=staff_instance)
    course_id_list = []
    for subject in subjects:
        course = Courses.objects.get(id=subject.course_id.id)
        course_id_list.append(course.id)

    final_course = []
    for course_id in course_id_list:
        if course_id not in final_course:
            final_course.append(course_id)

    students_count = Students.objects.filter(course_id__in=final_course).count()
    subject_count = subjects.count()

    attendance_count = Attendance.objects.filter(subject_id__in=subjects).count()

    subject_list = []
    attendance_list = []
    for subject in subjects:
        attendance_count1 = Attendance.objects.filter(subject_id=subject.id).count()
        subject_list.append(subject.subject_name)
        attendance_list.append(attendance_count1)

    # Get students and their attendance data
    students_attendance = Students.objects.filter(course_id__in=final_course)
    student_list = []
    student_list_attendance_present = []
    student_list_attendance_absent = []

    # Enhance students_attendance with attendance data
    for student in students_attendance:
        attendance_present_count = AttendanceReport.objects.filter(status=True, student_id=student.id).count()
        attendance_absent_count = AttendanceReport.objects.filter(status=False, student_id=student.id).count()
        # Add attendance data as attributes to student objects
        student.attendance_present = attendance_present_count
        student.attendance_absent = attendance_absent_count

        # Still maintain the lists for charts or other uses
        student_list.append(student.admin.first_name+" "+ student.admin.last_name)
        student_list_attendance_present.append(attendance_present_count)
        student_list_attendance_absent.append(attendance_absent_count)

    context={
        "students_count": students_count,
        "attendance_count": attendance_count,
        "subject_count": subject_count,
        "subject_list": subject_list,
        "attendance_list": attendance_list,
        "student_list": student_list,
        "attendance_present_list": student_list_attendance_present,
        "attendance_absent_list": student_list_attendance_absent,
        "students_attendance": students_attendance,
        "subjects": subjects
    }
    return render(request, "staff_template/staff_home_template.html", context)



from django.contrib.auth.decorators import login_required

@login_required
def staff_take_attendance(request):
    try:
        # Get the Staff instance linked to the logged-in user
        staff_instance = Staffs.objects.get(admin=request.user)

        # Fetch only subjects assigned to this staff
        subjects = Subjects.objects.filter(staff_id=staff_instance)

        # Check if a specific subject was requested
        selected_subject_id = request.GET.get('subject')
        selected_subject = None

        if selected_subject_id:
            try:
                selected_subject = subjects.get(id=selected_subject_id)
            except Subjects.DoesNotExist:
                # If the subject doesn't exist or doesn't belong to this staff
                pass

    except Staffs.DoesNotExist:
        # Handle case where staff entry is missing
        subjects = []
        selected_subject = None

    # Fetch all session years
    session_years = SessionYearModel.objects.all()

    context = {
        "subjects": subjects,
        "session_years": session_years,
        "selected_subject": selected_subject
    }
    return render(request, "staff_template/take_attendance_template.html", context)


@csrf_exempt
def get_students(request):
    subject_id = request.POST.get("subject")
    session_year = request.POST.get("session_year")

    # Validate input parameters
    if not subject_id or not session_year:
        return JsonResponse({"error": "Subject ID and Session Year are required"}, status=400)

    try:
        # Get the Staff instance linked to the logged-in user
        staff_instance = Staffs.objects.get(admin=request.user)

        # Verify the subject belongs to this staff
        subject_model = Subjects.objects.get(id=subject_id, staff_id=staff_instance)
        session_model = SessionYearModel.objects.get(id=session_year)

        # Get students for this course and session
        students = Students.objects.filter(course_id=subject_model.course_id, session_year_id=session_model)

        if students.count() == 0:
            # No students found for this course and session
            return JsonResponse({"message": "No students found for this course and session"}, safe=False)

    except Staffs.DoesNotExist:
        # Staff not found
        return JsonResponse({"error": "Staff not found"}, safe=False)
    except Subjects.DoesNotExist:
        # If subject doesn't exist or doesn't belong to this staff
        return JsonResponse({"error": "Subject not found or not assigned to you"}, safe=False)
    except SessionYearModel.DoesNotExist:
        # Session year not found
        return JsonResponse({"error": "Session year not found"}, safe=False)
    except Exception as e:
        # Other errors
        return JsonResponse({"error": str(e)}, safe=False)

    list_data = []

    for student in students:
        data_small={"id":student.admin.id, "name":student.admin.first_name+" "+student.admin.last_name}
        list_data.append(data_small)

    return JsonResponse(list_data, safe=False)




@csrf_exempt
def save_attendance_data(request):
    try:
        student_ids = request.POST.get("student_ids")
        subject_id = request.POST.get("subject_id")
        attendance_date = request.POST.get("attendance_date")
        session_year_id = request.POST.get("session_year_id")

        print(f"Save attendance data received:")
        print(f"- student_ids: {student_ids}")
        print(f"- subject_id: {subject_id}")
        print(f"- attendance_date: {attendance_date}")
        print(f"- session_year_id: {session_year_id}")

        if not all([student_ids, subject_id, attendance_date, session_year_id]):
            return HttpResponse("Error: Missing required fields")

        subject_model = Subjects.objects.get(id=subject_id)
        session_year_model = SessionYearModel.objects.get(id=session_year_id)

        json_student = json.loads(student_ids)
        print(f"Parsed student data: {json_student}")
    except Exception as e:
        print(f"Error in save_attendance_data setup: {str(e)}")
        return HttpResponse(f"Error: {str(e)}")

    try:
        # Check if attendance already exists for this date and subject
        existing_attendance = Attendance.objects.filter(
            subject_id=subject_model,
            attendance_date=attendance_date,
            session_year_id=session_year_model
        ).first()

        if existing_attendance:
            return HttpResponse("Error: Attendance already exists for this date and subject")

        attendance = Attendance(subject_id=subject_model, attendance_date=attendance_date, session_year_id=session_year_model)
        attendance.save()

        for stud in json_student:
            student = Students.objects.get(admin=stud['id'])

            # Check if attendance report already exists for this student and attendance
            existing_report = AttendanceReport.objects.filter(
                student_id=student,
                attendance_id=attendance
            ).first()

            if existing_report:
                # Update existing record instead of creating duplicate
                existing_report.status = stud['status']
                existing_report.location_verified = stud['status'] == 1
                existing_report.save()
            else:
                # Create new record
                location_verified = stud['status'] == 1  # True if present, False if absent
                attendance_report = AttendanceReport(
                    student_id=student,
                    attendance_id=attendance,
                    status=stud['status'],
                    location_verified=location_verified
                )
                attendance_report.save()
        return HttpResponse("OK")
    except Exception as e:
        return HttpResponse(f"Error: {str(e)}")




def staff_update_attendance(request):
    """Redirect to the combined manage attendance view"""
    from django.shortcuts import redirect
    return redirect('staff_view_attendance')

def staff_view_attendance(request):
    """Combined view for staff to view and update attendance records"""
    # Get the Staff instance linked to the logged-in user
    staff_instance = Staffs.objects.get(admin=request.user)

    # Fetch only subjects assigned to this staff
    subjects = Subjects.objects.filter(staff_id=staff_instance)

    # Fetch all session years
    session_years = SessionYearModel.objects.all()

    context = {
        "subjects": subjects,
        "session_years": session_years
    }
    return render(request, "staff_template/manage_attendance_template.html", context)

@csrf_exempt
def get_attendance_dates(request):
    # Get parameters from request
    subject_id = request.POST.get("subject_id")
    session_year = request.POST.get("session_year_id")

    try:
        # Validate parameters
        if not subject_id or not session_year:
            return JsonResponse(json.dumps([]), content_type="application/json", safe=False)

        # Get the Staff instance linked to the logged-in user
        staff_instance = Staffs.objects.get(admin=request.user)

        # Verify the subject belongs to this staff
        subject_model = Subjects.objects.get(id=subject_id, staff_id=staff_instance)
        session_model = SessionYearModel.objects.get(id=session_year)

        # Get attendance records for this subject and session
        attendance = Attendance.objects.filter(subject_id=subject_model, session_year_id=session_model)

    except (Subjects.DoesNotExist, SessionYearModel.DoesNotExist, Staffs.DoesNotExist, ValueError):
        # If any model doesn't exist or invalid data
        return JsonResponse([], safe=False)
    except Exception as e:
        # Log the error for debugging
        print(f"Error in get_attendance_dates: {str(e)}")
        return JsonResponse([], safe=False)

    list_data = []

    for attendance_single in attendance:
        # Format date properly for display
        formatted_date = attendance_single.attendance_date.strftime("%B %d, %Y")  # e.g., "January 15, 2024"
        data_small={
            "id": attendance_single.id,
            "attendance_date": formatted_date,
            "attendance_date_raw": attendance_single.attendance_date.strftime("%Y-%m-%d"),  # For calendar/sorting
            "session_year_id": attendance_single.session_year_id.id
        }
        list_data.append(data_small)

    return JsonResponse(list_data, safe=False)


@csrf_exempt
def get_attendance_student(request):
    try:
        attendance_id = request.POST.get('attendance_id')
        print(f"get_attendance_student called with attendance_id: {attendance_id}")

        if not attendance_id:
            print("No attendance_id provided")
            return JsonResponse([], safe=False)

        attendance = Attendance.objects.get(id=attendance_id)
        print(f"Found attendance record: {attendance}")
    except Attendance.DoesNotExist:
        print(f"Attendance with id {attendance_id} not found")
        return JsonResponse([], safe=False)
    except Exception as e:
        print(f"Error in get_attendance_student: {str(e)}")
        return JsonResponse([], safe=False)

    try:
        attendance_data = AttendanceReport.objects.filter(attendance_id=attendance)
        print(f"Found {attendance_data.count()} attendance records")

        list_data = []

        for student in attendance_data:
            try:
                # Include location_verified field in the response
                data_small={
                    "id": student.student_id.admin.id,
                    "name": student.student_id.admin.first_name+" "+student.student_id.admin.last_name,
                    "status": student.status,
                    "location_verified": student.location_verified
                }
                list_data.append(data_small)
            except Exception as e:
                print(f"Error processing student {student.id}: {str(e)}")
                continue

        print(f"Returning {len(list_data)} student records")
        return JsonResponse(list_data, safe=False)

    except Exception as e:
        print(f"Error getting attendance data: {str(e)}")
        return JsonResponse([], safe=False)


@csrf_exempt
def update_attendance_data(request):
    student_ids = request.POST.get("student_ids")

    attendance_date = request.POST.get("attendance_date")
    attendance = Attendance.objects.get(id=attendance_date)

    json_student = json.loads(student_ids)

    try:

        for stud in json_student:
            student = Students.objects.get(admin=stud['id'])

            attendance_report = AttendanceReport.objects.get(student_id=student, attendance_id=attendance)
            attendance_report.status = stud['status']

            # Update location_verified field based on status
            # If status is changing to present, set location_verified to True for manual attendance
            if stud['status'] == 1 and not attendance_report.status:
                attendance_report.location_verified = True
            # If status is changing to absent, set location_verified to False
            elif stud['status'] == 0:
                attendance_report.location_verified = False
            # Otherwise, keep the existing location_verified value

            attendance_report.save()
        return HttpResponse("OK")
    except:
        return HttpResponse("Error")


@csrf_exempt
def delete_attendance(request):
    """Delete an attendance record and all associated attendance reports"""
    if request.method != "POST":
        return JsonResponse({"status": "error", "message": "Invalid request method."}, status=400)

    attendance_id = request.POST.get("attendance_id")
    if not attendance_id:
        return JsonResponse({"status": "error", "message": "Attendance ID is required."}, status=400)

    try:
        # Get the attendance record
        attendance = Attendance.objects.get(id=attendance_id)

        # Check if the subject belongs to the staff
        if attendance.subject_id.staff_id.admin.id != request.user.id:
            return JsonResponse({"status": "error", "message": "You don't have permission to delete this attendance record."}, status=403)

        # Delete all attendance reports associated with this attendance
        AttendanceReport.objects.filter(attendance_id=attendance).delete()

        # Delete the attendance record
        attendance.delete()

        return JsonResponse({"status": "success", "message": "Attendance record deleted successfully."})

    except Attendance.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Attendance record not found."}, status=404)

    except Exception as e:
        return JsonResponse({"status": "error", "message": f"Error deleting attendance record: {str(e)}"}, status=500)


def staff_profile(request):
    user = CustomUser.objects.get(id=request.user.id)
    staff = Staffs.objects.get(admin=user)

    context={
        "user": user,
        "staff": staff
    }
    return render(request, 'staff_template/staff_profile.html', context)


def staff_profile_update(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method!")
        return redirect('staff_profile')
    else:
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        password = request.POST.get('password')
        address = request.POST.get('address')

        try:
            customuser = CustomUser.objects.get(id=request.user.id)
            customuser.first_name = first_name
            customuser.last_name = last_name
            if password != None and password != "":
                customuser.set_password(password)
            customuser.save()

            staff = Staffs.objects.get(admin=customuser.id)
            staff.address = address
            staff.save()

            messages.success(request, "Profile Updated Successfully")
            return redirect('staff_profile')
        except:
            messages.error(request, "Failed to Update Profile")
            return redirect('staff_profile')


def staff_apply_leave(request):
    """View for staff to apply for leave"""
    staff = Staffs.objects.get(admin=request.user.id)
    context = {
        "staff": staff
    }
    return render(request, "staff_template/staff_apply_leave.html", context)


def staff_feedback(request):
    """View for staff to provide feedback"""
    staff = Staffs.objects.get(admin=request.user.id)
    context = {
        "staff": staff
    }
    return render(request, "staff_template/staff_feedback.html", context)


def staff_add_result(request):
    # Get the Staff instance linked to the logged-in user
    staff_instance = Staffs.objects.get(admin=request.user)

    # Fetch only subjects assigned to this staff
    subjects = Subjects.objects.filter(staff_id=staff_instance)
    session_years = SessionYearModel.objects.all()
    context = {
        "subjects": subjects,
        "session_years": session_years,
    }
    return render(request, "staff_template/add_result_template.html", context)


def staff_add_result_save(request):
    if request.method != "POST":
        messages.error(request, "Invalid Method")
        return redirect('staff_add_result')
    else:
        student_admin_id = request.POST.get('student_list')
        assignment_marks = request.POST.get('assignment_marks')
        exam_marks = request.POST.get('exam_marks')
        subject_id = request.POST.get('subject')

        student_obj = Students.objects.get(admin=student_admin_id)
        subject_obj = Subjects.objects.get(id=subject_id)

        try:
            # Check if Students Result Already Exists or not
            check_exist = StudentResult.objects.filter(subject_id=subject_obj, student_id=student_obj).exists()
            if check_exist:
                result = StudentResult.objects.get(subject_id=subject_obj, student_id=student_obj)
                result.subject_assignment_marks = assignment_marks
                result.subject_exam_marks = exam_marks
                result.save()
                messages.success(request, "Result Updated Successfully!")
                return redirect('staff_add_result')
            else:
                result = StudentResult(student_id=student_obj, subject_id=subject_obj, subject_exam_marks=exam_marks, subject_assignment_marks=assignment_marks)
                result.save()
                messages.success(request, "Result Added Successfully!")
                return redirect('staff_add_result')
        except:
            messages.error(request, "Failed to Add Result!")
            return redirect('staff_add_result')

def staff_export_attendance(request):
    """View for exporting staff's attendance data"""
    # Get the staff ID from the user
    staff = Staffs.objects.get(admin=request.user.id)

    # Get subjects taught by this staff
    subjects = Subjects.objects.filter(staff_id=staff)

    # Get all session years
    session_years = SessionYearModel.objects.all()

    context = {
        "subjects": subjects,
        "session_years": session_years
    }
    return render(request, "staff_template/staff_export_attendance.html", context)


def staff_import_attendance(request):
    """View for staff to import attendance from Excel"""
    # Get the staff ID from the user
    staff = Staffs.objects.get(admin=request.user.id)

    # Get subjects taught by this staff
    subjects = Subjects.objects.filter(staff_id=staff)

    # Get all session years
    session_years = SessionYearModel.objects.all()

    # Get today's date for default value
    today_date = datetime.datetime.now().strftime('%Y-%m-%d')

    context = {
        "subjects": subjects,
        "session_years": session_years,
        "today_date": today_date,
        "pandas_available": PANDAS_AVAILABLE
    }
    return render(request, "staff_template/staff_import_attendance.html", context)


@csrf_exempt
def staff_import_attendance_data(request):
    """Process Excel file upload and import attendance data"""
    if request.method != 'POST':
        return JsonResponse({"status": "error", "message": "Invalid request method."})

    # Check if pandas is available
    if not PANDAS_AVAILABLE:
        return JsonResponse({
            "status": "error",
            "message": "The pandas library is not installed. Please install it using 'pip install pandas openpyxl' to use this feature."
        })

    try:
        subject_id = request.POST.get('subject')
        session_year_id = request.POST.get('session_year')
        attendance_date = request.POST.get('attendance_date')
        excel_file = request.FILES.get('excel_file')

        # Validate inputs
        if not subject_id or not session_year_id or not attendance_date or not excel_file:
            return JsonResponse({"status": "error", "message": "All fields are required."})

        # Check file extension
        if not excel_file.name.endswith(('.xls', '.xlsx')):
            return JsonResponse({"status": "error", "message": "Only Excel files (.xls, .xlsx) are allowed."})

        # Get subject and session year objects
        subject = Subjects.objects.get(id=subject_id)
        session_year = SessionYearModel.objects.get(id=session_year_id)

        # Check if subject belongs to the staff
        if subject.staff_id.admin.id != request.user.id:
            return JsonResponse({"status": "error", "message": "You don't have permission to import attendance for this subject."})

        # Parse Excel file - try to detect if it has title rows
        try:
            # First, try reading from row 0 (no title)
            df = pd.read_excel(excel_file, header=0)
            print(f"Excel columns (header=0): {df.columns.tolist()}")

            # Check if the first column looks like a title (contains spaces and is long)
            first_col = str(df.columns[0]).strip()
            if len(first_col) > 20 or 'attendance report' in first_col.lower():
                # This looks like a title row, try reading from row 3 or 4
                print("Detected title row, trying header=3")
                df = pd.read_excel(excel_file, header=3)
                print(f"Excel columns (header=3): {df.columns.tolist()}")

                # If still looks wrong, try header=4
                first_col = str(df.columns[0]).strip()
                if len(first_col) > 20 or 'attendance report' in first_col.lower():
                    print("Still looks like title, trying header=4")
                    df = pd.read_excel(excel_file, header=4)
                    print(f"Excel columns (header=4): {df.columns.tolist()}")
        except Exception as e:
            print(f"Error reading Excel: {e}")
            return JsonResponse({"status": "error", "message": f"Error reading Excel file: {str(e)}"})

        # Print final column names for debugging
        print(f"Final Excel columns: {df.columns.tolist()}")

        # Normalize column names (strip whitespace, convert to lowercase)
        df.columns = [str(col).strip().lower() for col in df.columns]

        # Check if required columns exist (case-insensitive)
        # Map expected column names to actual column names found in the file
        column_mapping = {}
        required_columns = ['student id', 'student name', 'status']

        for required_col in required_columns:
            found = False
            for actual_col in df.columns:
                if actual_col == required_col:
                    column_mapping[required_col] = actual_col
                    found = True
                    break
            if not found:
                return JsonResponse({"status": "error", "message": f"Column '{required_col}' is missing in the Excel file. Available columns: {', '.join(df.columns.tolist())}. Please use the export function to get the correct format."})

        # Check if Date column exists - it's optional but useful
        has_date_column = False
        date_column_name = None
        for col in df.columns:
            if col == 'date':
                has_date_column = True
                date_column_name = col
                break

        # Create or get attendance record
        attendance, _ = Attendance.objects.get_or_create(
            subject_id=subject,
            attendance_date=attendance_date,
            session_year_id=session_year
        )

        # Process each row in the Excel file
        success_count = 0
        error_count = 0
        processed_dates = set()

        for _, row in df.iterrows():
            try:
                # Get student ID from the Excel file using the mapped column name
                student_id_raw = row[column_mapping['student id']]

                # Skip empty rows or invalid data
                if pd.isna(student_id_raw) or str(student_id_raw).strip() == '' or str(student_id_raw).strip().lower() in ['nan', 'statistics', 'total records:', 'present count:', 'absent count:', 'location verified count:', 'location verification rate:']:
                    continue

                # Convert to string and strip any whitespace
                student_id = str(student_id_raw).strip()

                # Get date from the row if available, otherwise use the provided date
                row_date = None
                if has_date_column and date_column_name and not pd.isna(row[date_column_name]):
                    try:
                        # Try to parse the date from the Excel file
                        row_date = pd.to_datetime(row[date_column_name]).date()
                        processed_dates.add(row_date)
                    except:
                        # If date parsing fails, use the provided date
                        row_date = None

                # Use the date from the form if no date in the row or date parsing failed
                current_date = row_date or datetime.datetime.strptime(attendance_date, '%Y-%m-%d').date()

                # Handle different status formats (boolean, string, or numeric)
                status_value = row[column_mapping['status']]
                if isinstance(status_value, str):
                    # Handle string values like 'Present' or 'Absent'
                    status = status_value.lower() in ['present', 'yes', 'true', '1']
                elif isinstance(status_value, (int, float)):
                    # Handle numeric values (1 = Present, 0 = Absent)
                    status = bool(status_value)
                else:
                    # Default to boolean conversion
                    status = bool(status_value)

                # Get student object - prioritize username lookup since that's what we export
                try:
                    # First try to find by username (most common case)
                    custom_user = CustomUser.objects.get(username=str(student_id))
                    student = Students.objects.get(admin=custom_user)
                except (CustomUser.DoesNotExist, Students.DoesNotExist):
                    # If not found by username, try to find by admin_id (if it's a number)
                    try:
                        if str(student_id).isdigit():
                            student = Students.objects.get(admin_id=int(student_id))
                        else:
                            raise Exception(f"Student with username '{student_id}' not found")
                    except (Students.DoesNotExist, ValueError):
                        raise Exception(f"Student with ID '{student_id}' not found")

                # Get or create attendance record for this date
                attendance, attendance_created = Attendance.objects.get_or_create(
                    subject_id=subject,
                    attendance_date=current_date,
                    session_year_id=session_year
                )

                if attendance_created:
                    print(f"Created new attendance record for {subject.subject_name} on {current_date}")
                else:
                    print(f"Using existing attendance record for {subject.subject_name} on {current_date}")

                # Create or update attendance report
                attendance_report, created = AttendanceReport.objects.update_or_create(
                    student_id=student,
                    attendance_id=attendance,
                    defaults={'status': status}
                )

                # Log the operation for debugging
                action = "Created" if created else "Updated"
                print(f"{action} attendance for {student.admin.username} ({student.admin.first_name} {student.admin.last_name}) on {current_date}: {status}")
                print(f"  - Attendance ID: {attendance.id}, Student ID: {student.id}, Report ID: {attendance_report.id}")

                success_count += 1
            except Exception as e:
                print(f"Error processing row: {e}")
                error_count += 1
                continue

        # Create a more detailed success message
        dates_message = ""
        if processed_dates:
            date_list = sorted(processed_dates)
            if len(date_list) == 1:
                dates_message = f" for {date_list[0]}"
            elif len(date_list) > 1:
                dates_message = f" for {len(date_list)} dates from {date_list[0]} to {date_list[-1]}"

        # Verify the import by checking if records exist
        verification_message = ""
        if success_count > 0:
            try:
                # Check how many attendance reports exist for this subject and date range
                if processed_dates:
                    total_reports = AttendanceReport.objects.filter(
                        attendance_id__subject_id=subject,
                        attendance_id__attendance_date__in=processed_dates
                    ).count()
                else:
                    import_date = datetime.datetime.strptime(attendance_date, '%Y-%m-%d').date()
                    total_reports = AttendanceReport.objects.filter(
                        attendance_id__subject_id=subject,
                        attendance_id__attendance_date=import_date
                    ).count()
                verification_message = f" Verification: {total_reports} total attendance records found in database."
            except Exception as e:
                verification_message = f" Verification failed: {str(e)}"

        return JsonResponse({
            "status": "success",
            "message": f"Attendance imported successfully{dates_message}. {success_count} records processed, {error_count} errors.{verification_message}"
        })

    except Subjects.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Subject not found."})
    except SessionYearModel.DoesNotExist:
        return JsonResponse({"status": "error", "message": "Session year not found."})
    except Exception as e:
        return JsonResponse({"status": "error", "message": f"Error importing attendance: {str(e)}"})


def staff_download_import_template(request):
    """Download a blank Excel template for attendance import"""
    try:
        # Create a new workbook
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Attendance Import Template"

        # Add headers that match exactly what the import function expects
        headers = ['Student ID', 'Student Name', 'Date', 'Status']
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)

        # Add sample data rows
        sample_data = [
            ['student1', 'John Doe', '2024-01-15', 'Present'],
            ['student2', 'Jane Smith', '2024-01-15', 'Absent'],
            ['student3', 'Bob Johnson', '2024-01-15', 'Present']
        ]

        for row_num, row_data in enumerate(sample_data, 2):
            for col_num, value in enumerate(row_data, 1):
                worksheet.cell(row=row_num, column=col_num).value = value

        # Add instructions
        worksheet.cell(row=6, column=1).value = "Instructions:"
        worksheet.cell(row=6, column=1).font = openpyxl.styles.Font(bold=True)
        worksheet.cell(row=7, column=1).value = "1. Replace sample data with actual student data"
        worksheet.cell(row=8, column=1).value = "2. Student ID must match the username in the system"
        worksheet.cell(row=9, column=1).value = "3. Date format: YYYY-MM-DD (optional)"
        worksheet.cell(row=10, column=1).value = "4. Status must be 'Present' or 'Absent'"
        worksheet.cell(row=11, column=1).value = "5. Do not change the column headers"

        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        workbook.save(temp_file.name)

        # Read file content
        with open(temp_file.name, 'rb') as f:
            file_content = f.read()

        # Clean up
        os.unlink(temp_file.name)

        # Return file response
        response = HttpResponse(file_content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="attendance_import_template.xlsx"'
        return response

    except Exception as e:
        return JsonResponse({"status": "error", "message": f"Error generating template: {str(e)}"})


@csrf_exempt
def staff_export_attendance_data(request):
    if request.method != 'POST':
        return HttpResponse("Method Not Allowed", status=405)

    subject_id = request.POST.get('subject')
    session_year_id = request.POST.get('session_year')
    start_date = request.POST.get('start_date')
    end_date = request.POST.get('end_date')

    if not subject_id or not session_year_id or not start_date or not end_date:
        return JsonResponse({'status': 'error', 'message': 'Missing required fields'})

    try:
        subject = Subjects.objects.get(id=subject_id)
        session_year = SessionYearModel.objects.get(id=session_year_id)

        # Convert string dates to datetime objects
        start_date = datetime.datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.datetime.strptime(end_date, '%Y-%m-%d').date()

        # Get attendance records for the specified period
        attendances = Attendance.objects.filter(
            subject_id=subject,
            session_year_id=session_year,
            attendance_date__range=(start_date, end_date)
        )

        if not attendances:
            # Check if there are any attendance records for this subject at all
            any_attendance = Attendance.objects.filter(subject_id=subject).exists()

            if any_attendance:
                return JsonResponse({
                    'status': 'error',
                    'message': f'No attendance records found between {start_date} and {end_date} for {subject.subject_name}. Please try a different date range.'
                })
            else:
                return JsonResponse({
                    'status': 'error',
                    'message': f'No attendance records found for {subject.subject_name}. Please take attendance first before exporting.'
                })

        # Get all attendance reports for these attendances
        attendance_reports = AttendanceReport.objects.filter(attendance_id__in=attendances)

        if not attendance_reports:
            return JsonResponse({'status': 'error', 'message': 'No attendance data found for the selected criteria'})

        # Generate Excel file
        date_range = f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
        excel_file = export_attendance_to_excel(
            attendance_reports,
            subject_name=subject.subject_name,
            date_range=date_range
        )

        # Prepare filename
        filename = f"attendance_{subject.subject_name}_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"

        # Read the file content instead of keeping it open
        with open(excel_file, 'rb') as f:
            file_content = f.read()

        # Create the response with the content
        response = HttpResponse(
            file_content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # Set headers to force download
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = len(file_content)

        # Close and delete the file explicitly before returning the response
        try:
            os.unlink(excel_file)
        except Exception as file_error:
            # If we can't delete now, schedule it for later
            import atexit
            atexit.register(lambda file_path=excel_file: os.unlink(file_path) if os.path.exists(file_path) else None)
            print(f"Warning: Could not delete temp file immediately: {str(file_error)}")

        return response

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'status': 'error', 'message': str(e)})


# Network verification function removed
